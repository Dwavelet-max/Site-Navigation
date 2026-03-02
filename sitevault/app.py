import os
import uuid
import shutil
from flask import Flask, render_template, jsonify, request, send_from_directory, session, redirect, url_for, send_file
from flask_session import Session
from sitevault.models import db, Site, Category, User
from sitevault.snapshot import save_html_snapshot
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sitevault.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
app.config['SESSION_TYPE'] = 'filesystem'

db.init_app(app)
Session(app)

# 创建数据库表和默认数据
with app.app_context():
    db.create_all()
    
    if not User.query.first():
        admin = User(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        
        default_cats = ['技术', '资讯', '工具', '学习', '其他']
        for name in default_cats:
            db.session.add(Category(user_id=admin.id, name=name))
        db.session.commit()


def get_current_user():
    user_id = session.get('user_id')
    if user_id:
        return User.query.get(user_id)
    return None


@app.route('/')
def index():
    if not session.get('user_id'):
        return render_template('index.html', logged_in=False)
    return render_template('index.html', logged_in=True)


@app.route('/api/auth/check', methods=['GET'])
def check_auth():
    if session.get('user_id'):
        user = User.query.get(session['user_id'])
        return jsonify({'logged_in': True, 'username': user.username})
    return jsonify({'logged_in': False})


@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    
    if len(password) < 4:
        return jsonify({'error': '密码至少4位'}), 400
    
    if User.query.filter_by(username=username).first():
        return jsonify({'error': '用户名已存在'}), 400
    
    user = User(username=username)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    
    session['user_id'] = user.id
    return jsonify({'success': True, 'username': user.username})


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    user = User.query.filter_by(username=username).first()
    if not user or not user.check_password(password):
        return jsonify({'error': '用户名或密码错误'}), 401
    
    session['user_id'] = user.id
    return jsonify({'success': True, 'username': user.username})


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/password', methods=['PUT'])
def change_password():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    data = request.json
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    
    if not old_password or not new_password:
        return jsonify({'error': '密码不能为空'}), 400
    
    if len(new_password) < 4:
        return jsonify({'error': '新密码至少4位'}), 400
    
    user = User.query.get(session['user_id'])
    if not user.check_password(old_password):
        return jsonify({'error': '原密码错误'}), 400
    
    user.set_password(new_password)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/auth/user', methods=['GET'])
def get_user_info():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    user = User.query.get(session['user_id'])
    return jsonify({
        'id': user.id,
        'username': user.username,
        'created_at': user.created_at.strftime('%Y-%m-%d') if user.created_at else ''
    })


@app.route('/api/sites', methods=['GET'])
def get_sites():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    category_id = request.args.get('category_id', type=int)
    search = request.args.get('search', '')
    
    query = Site.query.filter_by(user_id=session['user_id'])
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(Site.title.contains(search) | Site.url.contains(search))
    
    sites = query.order_by(Site.updated_at.desc()).all()
    return jsonify([s.to_dict() for s in sites])


@app.route('/api/sites', methods=['POST'])
def add_site():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    data = request.json
    site = Site(
        user_id=session['user_id'],
        url=data['url'],
        title=data.get('title', data['url']),
        category_id=data.get('category_id'),
        note=data.get('note', '')
    )
    db.session.add(site)
    db.session.commit()
    return jsonify(site.to_dict())


@app.route('/api/sites/<int:id>', methods=['PUT'])
def update_site(id):
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    site = Site.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    data = request.json
    site.url = data.get('url', site.url)
    site.title = data.get('title', site.title)
    site.category_id = data.get('category_id', site.category_id)
    site.note = data.get('note', site.note)
    db.session.commit()
    return jsonify(site.to_dict())


@app.route('/api/sites/<int:id>', methods=['DELETE'])
def delete_site(id):
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    site = Site.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if site.snapshot_path:
        snapshot_dir = os.path.join(app.config['UPLOAD_FOLDER'], f"{session['user_id']}_{site.snapshot_path}")
        if os.path.exists(snapshot_dir):
            shutil.rmtree(snapshot_dir)
    if site.screenshot_path:
        screenshot_file = os.path.join(app.config['UPLOAD_FOLDER'], f"{session['user_id']}_{site.screenshot_path}")
        if os.path.exists(screenshot_file):
            os.remove(screenshot_file)
    
    db.session.delete(site)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/sites/<int:id>/snapshot', methods=['POST'])
def save_snapshot(id):
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    site = Site.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if site.snapshot_path:
        old_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session['user_id']}_{site.snapshot_path}")
        if os.path.exists(old_path):
            shutil.rmtree(old_path)
    
    folder_name = save_html_snapshot(site.url, app.config['UPLOAD_FOLDER'], session['user_id'])
    
    if folder_name:
        site.snapshot_path = folder_name
        db.session.commit()
        return jsonify({'success': True, 'path': folder_name})
    return jsonify({'success': False, 'error': '保存失败'}), 400


@app.route('/api/sites/<int:id>/view')
def view_snapshot(id):
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    site = Site.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    if not site.snapshot_path:
        return '无离线内容', 404
    
    snapshot_dir = os.path.join(app.config['UPLOAD_FOLDER'], site.snapshot_path)
    return send_from_directory(snapshot_dir, 'index.html')


@app.route('/api/categories', methods=['GET'])
def get_categories():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    categories = Category.query.filter_by(user_id=session['user_id']).all()
    return jsonify([c.to_dict() for c in categories])


@app.route('/api/categories', methods=['POST'])
def add_category():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    data = request.json
    if Category.query.filter_by(user_id=session['user_id'], name=data['name']).first():
        return jsonify({'error': '分类已存在'}), 400
    category = Category(user_id=session['user_id'], name=data['name'])
    db.session.add(category)
    db.session.commit()
    return jsonify(category.to_dict())


@app.route('/api/categories/<int:id>', methods=['DELETE'])
def delete_category(id):
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    category = Category.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    Site.query.filter_by(category_id=id, user_id=session['user_id']).update({'category_id': None})
    db.session.delete(category)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/stats')
def get_stats():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    total = Site.query.filter_by(user_id=user_id).count()
    with_snapshot = Site.query.filter(Site.user_id == user_id, Site.snapshot_path != '').count()
    categories = Category.query.filter_by(user_id=user_id).all()
    
    cat_data = []
    for cat in categories:
        cat_data.append({'name': cat.name, 'value': len(cat.sites)})
    
    uncategorized = Site.query.filter(Site.user_id == user_id, Site.category_id == None).count()
    if uncategorized > 0:
        cat_data.append({'name': '未分类', 'value': uncategorized})
    
    return jsonify({
        'total': total,
        'with_snapshot': with_snapshot,
        'categories': len(categories),
        'category_data': cat_data,
        'snapshot_rate': round(with_snapshot / total * 100, 1) if total > 0 else 0
    })


@app.route('/api/sites/export')
def export_sites():
    if not session.get('user_id'):
        return jsonify({'error': '未登录'}), 401
    
    try:
        user_id = session['user_id']
        sites = Site.query.filter_by(user_id=user_id).order_by(Site.updated_at.desc()).all()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '站点列表'
        
        # 设置表头
        headers = ['标题', 'URL', '分类', '备注', '创建时间', '更新时间', '是否有快照']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for row, site in enumerate(sites, 2):
            ws.cell(row=row, column=1).value = site.title
            ws.cell(row=row, column=2).value = site.url
            ws.cell(row=row, column=3).value = site.category.name if site.category else '未分类'
            ws.cell(row=row, column=4).value = site.note
            ws.cell(row=row, column=5).value = site.created_at.strftime('%Y-%m-%d %H:%M') if site.created_at else ''
            ws.cell(row=row, column=6).value = site.updated_at.strftime('%Y-%m-%d %H:%M') if site.updated_at else ''
            ws.cell(row=row, column=7).value = '是' if site.snapshot_path else '否'
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].auto_size = True
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 发送文件
        return send_file(output, as_attachment=True, download_name='sitevault_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--hostname', default='0.0.0.0')
    parser.add_argument('--port', type=int, default=5000)
    args = parser.parse_args()
    
    with app.app_context():
        db.create_all()
        
        from flask_session import Session
        Session(app)
        
        if not User.query.first():
            admin = User(username='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            
            default_cats = ['技术', '资讯', '工具', '学习', '其他']
            user = User.query.first()
            for name in default_cats:
                db.session.add(Category(user_id=user.id, name=name))
            db.session.commit()
    
    app.run(debug=True, port=args.port, host=args.hostname)
